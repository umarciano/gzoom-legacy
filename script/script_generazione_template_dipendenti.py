#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Genera `Template_Dipendenti_AORN.xlsx` a partire da DIPENDENTI_2025_BASE_GOP_LAVORATO.xlsx.

File di input : <script_dir>/templates/DIPENDENTI_2025_BASE_GOP_LAVORATO.xlsx
                Sheet: "Lista dipendenti"
File di output: <script_dir>/templates/Template_Dipendenti_AORN.xlsx
                Sheet: "Template Dipendenti AORN"

Mapping colonne input -> output:
  MATRICOLA            -> Matricola
  Cognome              -> Cognome
  Nome                 -> Nome
  CF                   -> Codice Fiscale
  DESC QUALIFICA       -> Descrizione INCARICHI ECONOMICI
  CdC_new              -> Codice UOC
  Periodo completo     -> Decorrenza
  Periodo completo FINE-> Scadenza
  Mail                 -> Email  (con logica cessato, vedi sotto)
  Mt. valutatore       -> Matricola Referente Valutatore

Valori di default:
  Codifica             -> 'd'
  Descrizione ESCLUSIVO-> 'RAPPORTO ESCLUSIVO'
  Username             -> parte della mail prima del '@'

Logica Mail/Email:
  - Se Mail contiene 'cessato' (case-insensitive):
      email = nome.cognome@aocardarelli.it  (costruito da Nome e Cognome normalizzati)
  - Altrimenti:
      email = valore della colonna Mail as-is

Username = email senza dominio (parte prima di '@').
"""

import sys
from pathlib import Path
import re
import unicodedata
import pandas as pd
from datetime import datetime


CONFIG = {
    'TEMPLATE_DIR': 'templates',
    'INPUT_FILE':   'DIPENDENTI_2025_BASE_GOP_LAVORATO.xlsx',
    'INPUT_SHEET':  'Lista dipendenti',
    'OUTPUT_FILE':  'Template_Dipendenti_AORN.xlsx',
    'OUTPUT_SHEET': 'Template Dipendenti AORN',
    'EMAIL_DOMAIN': 'aocardarelli.it',
}


def clean_value(v):
    if pd.isna(v):
        return ''
    s = str(v).strip()
    if s.endswith('.0'):
        try:
            float(s)
            return s[:-2]
        except ValueError:
            return s
    return s


def format_date(v):
    if pd.isna(v) or v == '':
        return ''
    if isinstance(v, datetime):
        return v.strftime('%d/%m/%Y')
    # try parse
    s = str(v).strip()
    # if already in dd/mm/yyyy-ish, return
    return s


def _username_part(s: str) -> str:
    """Normalize a name part for username: remove accents, spaces, punctuation, lowercase."""
    if not s:
        return ''
    s = clean_value(s)
    if not s:
        return ''
    # remove accents
    s = unicodedata.normalize('NFKD', s)
    s = s.encode('ascii', 'ignore').decode('ascii')
    s = s.lower()
    # remove spaces and non-alphanumeric
    s = re.sub(r'\s+', '', s)
    s = re.sub(r'[^a-z0-9]', '', s)
    return s


def _sanitize_for_excel(s: str) -> str:
    """Prevent Excel from interpreting strings as formulas by prefixing a single quote
    when the string starts with characters that can trigger formula parsing.
    """
    if s is None:
        return ''
    try:
        if not isinstance(s, str):
            s = str(s)
    except Exception:
        s = ''
    if s == '':
        return s
    # If string starts with '=' or '-' remove those leading chars (user requested)
    s = s.lstrip()
    s = re.sub(r'^[=\-]+', '', s)
    return s


def _find_col(df_columns, *candidates):
    """Restituisce il nome della colonna effettiva (case-insensitive) tra i candidati."""
    lower_map = {c.lower(): c for c in df_columns if c is not None}
    for cand in candidates:
        found = lower_map.get(cand.lower())
        if found is not None:
            return found
    return None


def _get(row, df_columns, *candidates):
    """Legge il valore di riga dalla prima colonna candidata trovata (case-insensitive)."""
    col = _find_col(df_columns, *candidates)
    if col is None:
        return ''
    return clean_value(row.get(col))


def main():
    script_dir = Path(__file__).parent
    template_dir = script_dir / CONFIG['TEMPLATE_DIR']
    template_dir.mkdir(parents=True, exist_ok=True)

    input_file  = template_dir / CONFIG['INPUT_FILE']
    output_file = template_dir / CONFIG['OUTPUT_FILE']

    # Supporto argomenti opzionali per override
    if len(sys.argv) >= 2 and sys.argv[1].strip():
        input_file = Path(sys.argv[1])
    if len(sys.argv) >= 3 and sys.argv[2].strip():
        output_file = Path(sys.argv[2])

    if not input_file.exists():
        print(f"Errore: file di input {input_file} non trovato")
        sys.exit(1)

    print(f"Input : {input_file}")
    print(f"Output: {output_file}")

    # Leggi lo sheet "Lista dipendenti"
    df_in = pd.read_excel(input_file, sheet_name=CONFIG['INPUT_SHEET'])
    cols  = list(df_in.columns)

    rows = []

    for _, r in df_in.iterrows():
        # Salta righe senza matricola
        matricola = _get(r, cols, 'MATRICOLA', 'Matricola', 'matricola')
        if matricola == '':
            continue

        cognome       = _get(r, cols, 'Cognome', 'COGNOME', 'cognome')
        nome          = _get(r, cols, 'Nome', 'NOME', 'nome')
        cf            = _get(r, cols, 'CF', 'Codice Fiscale')
        desc_qualifica= _get(r, cols, 'DESC QUALIFICA', 'Desc Qualifica', 'DESC_QUALIFICA')
        codice_uoc    = _get(r, cols, 'CdC_new', 'CDC_NEW', 'cdc_new')
        mat_val       = _get(r, cols, 'Mt. valutatore', 'MT. VALUTATORE', 'Mt valutatore', 'Mt_valutatore')

        # Campi lasciati vuoti (TODO: COMPLETARE)
        ruolo_gzoom = ''
        tipo_scheda = ''
        nome_uoc    = ''
        codice_dip  = ''
        nome_dip    = ''

        descrizione_esclusivo = 'RAPPORTO ESCLUSIVO'

        # Decorrenza / Scadenza
        col_dec = _find_col(cols, 'Periodo completo', 'DATA INIZIO', 'Data Inizio')
        col_sca = _find_col(cols, 'Periodo completo FINE', 'DATA FINE', 'Data Fine')
        decorrenza = format_date(r.get(col_dec)) if col_dec else ''
        scadenza   = format_date(r.get(col_sca)) if col_sca else ''

        data_nascita = ''

        # ------------------------------------------------------------------
        # Email / Username
        #   Se Mail contiene 'cessato' (case-insensitive):
        #       email = nome.cognome@aocardarelli.it  (da Nome e Cognome)
        #   Altrimenti:
        #       email = valore Mail as-is
        #   Username = email senza dominio (parte prima di '@')
        # ------------------------------------------------------------------
        mail_raw = _get(r, cols, 'Mail', 'MAIL', 'mail', 'Email', 'EMAIL')

        if 'cessato' in mail_raw.lower():
            part_nome    = _username_part(nome)
            part_cognome = _username_part(cognome)
            local = f"{part_nome}.{part_cognome}" if part_nome and part_cognome else (part_nome or part_cognome)
            email = f"{local}@{CONFIG['EMAIL_DOMAIN']}" if local else ''
        else:
            email = mail_raw

        username = email.split('@')[0] if '@' in email else email

        row_out = {
            'Codifica'                        : 'd',
            'Matricola'                       : matricola,
            'Cognome'                         : cognome,
            'Nome'                            : nome,
            'Codice Fiscale'                  : cf,
            'Descrizione INCARICHI ECONOMICI' : desc_qualifica,
            'Ruolo GZOOM'                     : ruolo_gzoom,
            'Tipo Scheda'                     : tipo_scheda,
            'Codice UOC'                      : codice_uoc,
            'Nome UOC'                        : nome_uoc,
            'Codice Dipartimento'             : codice_dip,
            'Nome Dipartimento'               : nome_dip,
            'Descrizione ESCLUSIVO'           : descrizione_esclusivo,
            'Decorrenza'                      : decorrenza,
            'Scadenza'                        : scadenza,
            'Data di Nascita'                 : data_nascita,
            'Email'                           : email,
            'Username'                        : username,
            'Matricola Referente Valutatore'  : mat_val,
        }

        rows.append(row_out)

    if not rows:
        print('Nessuna riga valida trovata nel file di input.')
        sys.exit(0)

    df_out = pd.DataFrame(rows)

    # Forza tutte le colonne a stringa
    for c in df_out.columns:
        df_out[c] = df_out[c].astype(str).replace('nan', '')

    # Sanitizza celle (evita formule accidentali)
    df_out = df_out.applymap(lambda x: _sanitize_for_excel(x))

    # Scrivi preservando formule/stile del file esistente con openpyxl
    import openpyxl
    from openpyxl import load_workbook

    def _write_output(path):
        if path.exists():
            wb = load_workbook(path)
        else:
            wb = openpyxl.Workbook()

        sheet_name = CONFIG['OUTPUT_SHEET']
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Cancella solo le righe dati (dalla riga 2 in poi), preserva header e stile
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.value = None
        else:
            ws = wb.create_sheet(sheet_name)
            # Scrivi header
            for ci, col_name in enumerate(df_out.columns, start=1):
                ws.cell(row=1, column=ci, value=col_name)

        # Scrivi dati dalla riga 2
        for ri, row_data in enumerate(df_out.itertuples(index=False), start=2):
            for ci, val in enumerate(row_data, start=1):
                ws.cell(row=ri, column=ci, value=val if val != '' else None)

        wb.save(path)
        print(f"Generato file: {path} (righe: {len(df_out)})")

    try:
        _write_output(output_file)
    except PermissionError:
        fallback = output_file.with_name(output_file.stem + '_generated' + output_file.suffix)
        try:
            _write_output(fallback)
            print(f"File originale bloccato. Generato file alternativo: {fallback}")
        except Exception as e:
            print(f"Errore scrittura file di output: {e}")
            raise


if __name__ == '__main__':
    main()
